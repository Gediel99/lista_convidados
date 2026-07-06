import csv
import io
import json
import html
import urllib.parse
import uuid
from pathlib import Path

import streamlit as st

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


APP_DIR = Path(__file__).parent
DATA_FILE = APP_DIR / "dados_casamento.json"
DEFAULT_XLSX = Path(r"C:\Users\gedie\OneDrive\Desktop\Pasta1.xlsx")

INITIAL_FAMILIES = [
    ["Sebastião", "Maria", "Elizeu", "Raquel", "Lucas"],
    ["Alessandro", "Jaqueline", "Diogo", "Eliza", "Nelis", "Vilmar"],
    ["Alcebiados", "Zelita"],
    ["Wesli"],
    ["Marcos", "Elci", "Geise"],
    ["Micael", "Manuele", "Elói"],
    ["Dora", "Luiza", "Vitor"],
    ["Valdir", "Roseli", "Erica"],
    ["Elder", "Priscila", "Alice"],
    ["José"],
    ["Amanda", "Alice", "Mirtes", "Gentil"],
    ["Marcolina", "Beto"],
    ["Eduardo", "Mari", "Heitor"],
    ["Olavo", "Elizangela", "Jonatan"],
    ["Filho", "Ana", "Jonatan", "Evilyn"],
    ["Victor", "Fledson", "Neia"],
    ["Eliel", "Monalisa", "Isaque"],
    ["Marcelo", "Juliana", "Antony"],
    ["Adriano", "Manuele"],
    ["Gesiel", "Paula", "Wendril Gabriel"],
]


# -----------------------------
# Dados
# -----------------------------

def family_label(index, names):
    first_name = names[0] if names else f"Família {index}"
    return f"Família {first_name}"


def display_family_name(family):
    family = (family or "Sem família").strip()
    if family.startswith("Familia "):
        return "Família " + family[len("Familia "):]
    if family == "Sem familia":
        return "Sem família"
    return family


def new_guest(name="", family="", favorite=False):
    return {
        "id": str(uuid.uuid4()),
        "name": name,
        "family": display_family_name(family or "Sem família"),
        "favorite": favorite,
        "status": "Pendente",
        "side": "Noiva",
    }


def default_data():
    guests = []
    for index, names in enumerate(INITIAL_FAMILIES, start=1):
        family = family_label(index, names)
        guests.extend(new_guest(name, family) for name in names)
    return {"guests": guests}


def normalize_guest(item):
    if "role" in item:
        favorite = item.get("role") in ["Padrinho", "Madrinha"]
    else:
        favorite = bool(item.get("favorite", False))

    return {
        "id": item.get("id", str(uuid.uuid4())),
        "name": item.get("name", ""),
        "family": display_family_name(item.get("family", "Sem família")),
        "favorite": favorite,
        "status": item.get("status", "Pendente"),
        "side": item.get("side", "Noiva"),
    }


def load_data():
    if DATA_FILE.exists():
        try:
            data = json.loads(DATA_FILE.read_text(encoding="utf-8"))
            if "guests" in data:
                return {"guests": [normalize_guest(g) for g in data["guests"]]}
            if "people" in data:
                return {"guests": [normalize_guest(g) for g in data["people"]]}
        except json.JSONDecodeError:
            pass
    return default_data()


def save_data():
    DATA_FILE.write_text(
        json.dumps({"guests": st.session_state.guests}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def update_guest(guest_id, field, value):
    for guest in st.session_state.guests:
        if guest["id"] == guest_id:
            guest[field] = value
            break
    save_data()


def remove_guest(guest_id):
    st.session_state.guests = [
        guest for guest in st.session_state.guests if guest["id"] != guest_id
    ]
    if st.session_state.get("selected_guest_id") == guest_id:
        st.session_state.selected_guest_id = None
    if st.session_state.get("editing_guest_id") == guest_id:
        st.session_state.editing_guest_id = None
    save_data()


def add_guest_from_inputs():
    name = st.session_state.get("new_guest_name", "").strip()
    family = st.session_state.get("new_guest_family", "").strip() or "Sem família"
    side = st.session_state.get("new_guest_side", "Noiva")
    if not name:
        return
    guest = new_guest(name, family)
    guest["side"] = side
    st.session_state.guests.insert(0, guest)
    st.session_state.editing_guest_id = guest["id"]
    st.session_state.selected_guest_id = guest["id"]
    st.session_state.new_guest_name = ""
    save_data()


def add_guest_to_family(family, input_key):
    name = st.session_state.get(input_key, "").strip()
    if not name:
        return
    guest = new_guest(name, display_family_name(family))
    st.session_state.guests.insert(0, guest)
    st.session_state.editing_guest_id = guest["id"]
    st.session_state.selected_guest_id = guest["id"]
    st.session_state[input_key] = ""
    save_data()


def rename_family(old_family, new_family):
    new_family = display_family_name((new_family or "").strip() or "Sem família")
    for guest in st.session_state.guests:
        if family_matches(guest, old_family):
            guest["family"] = new_family
    st.session_state.selected_family = new_family
    save_data()


def import_xlsx(path):
    if load_workbook is None:
        st.error("Instale openpyxl para importar Excel.")
        return

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    guests = []

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        names = []
        favorites = []
        for index, value in enumerate(row):
            if value is None:
                continue
            text = str(value).strip()
            if not text or text.isdigit() or text.lower().startswith("possiveis"):
                continue
            if index == 7:
                favorites.append(text)
            elif index < 7:
                names.append(text)

        if names:
            family = family_label(row_number, names)
            guests.extend(new_guest(name, family) for name in names)
        for name in favorites:
            guests.append(new_guest(name, "Favoritos da planilha", True))

    st.session_state.guests = guests
    save_data()


def csv_bytes():
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["nome", "familia", "favorito", "status", "lado"])
    for guest in st.session_state.guests:
        writer.writerow([
            guest["name"],
            guest["family"],
            "sim" if guest["favorite"] else "nao",
            guest.get("status", "Pendente"),
            guest.get("side", "Noiva"),
        ])
    return buffer.getvalue().encode("utf-8-sig")


def family_summary():
    families = {}
    for guest in st.session_state.guests:
        if not guest["name"].strip():
            continue
        family = display_family_name(guest["family"].strip() or "Sem família")
        families.setdefault(family, {"total": 0, "favorites": 0})
        families[family]["total"] += 1
        families[family]["favorites"] += int(guest["favorite"])
    return families


def stable_key(value):
    return uuid.uuid5(uuid.NAMESPACE_URL, str(value)).hex


def family_matches(guest, family):
    current = display_family_name((guest.get("family") or "Sem família").strip() or "Sem família")
    return current == display_family_name(family)


def find_guest(guest_id):
    return next((guest for guest in st.session_state.guests if guest["id"] == guest_id), None)


# -----------------------------
# Navegação / query params
# -----------------------------

def qp_value(name, default=None):
    value = st.query_params.get(name, default)
    if isinstance(value, list):
        return value[0] if value else default
    return value


def build_url(page, **params):
    query = {"page": page}
    for key, value in params.items():
        if value is not None:
            query[key] = value
    return "?" + urllib.parse.urlencode(query)


def set_page(page):
    st.session_state.page = page
    st.session_state.selected_guest_id = None
    st.session_state.editing_guest_id = None
    st.session_state.selected_family = None
    st.query_params.clear()
    st.query_params["page"] = page


def clear_detail_params():
    st.session_state.selected_guest_id = None
    st.session_state.editing_guest_id = None
    st.session_state.selected_family = None
    current_page = st.session_state.page
    st.query_params.clear()
    st.query_params["page"] = current_page


# -----------------------------
# Componentes visuais
# -----------------------------

def render_header(title, subtitle, icon):
    st.markdown(
        f"""
        <div class="app-header">
            <div>
                <h1 class="app-title">{html.escape(title)}</h1>
                <div class="app-subtitle">{html.escape(subtitle)}</div>
            </div>
            <div class="header-icon">{icon}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_family_stats(family_count, total_guests):
    st.markdown(
        f"""
        <div class="stats-card">
            <div class="stats-item">
                <div class="stats-icon">👥</div>
                <div>
                    <div class="stats-number">{family_count}</div>
                    <div class="stats-label">famílias</div>
                </div>
            </div>
            <div class="stats-divider"></div>
            <div class="stats-item">
                <div class="stats-icon">♡</div>
                <div>
                    <div class="stats-number">{total_guests}</div>
                    <div class="stats-label">convidados</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_guest_card(guest, page="Convidados", family_locked=False):
    """Renderiza o convidado como card clicável e expande a edição no próprio card.

    Não usa link, query string ou nova tela. O clique apenas altera o estado da sessão
    e abre o formulário logo abaixo do card selecionado.
    """
    name_raw = guest.get("name", "").strip() or "Sem nome"
    name = html.escape(name_raw)
    family = html.escape(display_family_name(guest.get("family", "Sem família")))
    is_open = st.session_state.get("editing_guest_id") == guest["id"]
    favorite_badge = '<span class="mini-badge">Padrinho</span>' if guest.get("favorite") else ""
    open_class = " guest-card-open" if is_open else ""
    chevron = "⌄" if is_open else "›"

    card_html = (
        f'<div class="guest-card{open_class}">'
        '<div class="guest-avatar">♡</div>'
        '<div class="guest-copy">'
        f'<div class="guest-name">{name}</div>'
        f'<div class="guest-family">{family}</div>'
        '</div>'
        f'{favorite_badge}'
        f'<div class="guest-chevron">{chevron}</div>'
        '</div>'
        '<span class="guest-card-click-marker"></span>'
    )
    st.markdown(card_html, unsafe_allow_html=True)

    button_label = "Fechar edição" if is_open else f"Editar {name_raw}"
    if st.button(button_label, key=f"open_guest_{page}_{guest['id']}", use_container_width=True):
        st.session_state.editing_guest_id = None if is_open else guest["id"]
        st.session_state.selected_guest_id = None if is_open else guest["id"]
        st.rerun()

    if is_open:
        render_guest_editor_panel(guest, "Editar convidado", family_locked=family_locked)


def render_family_row(family, values):
    """Card de família clicável sem link/aba nova."""
    family_display = display_family_name(family)
    safe_family = html.escape(family_display)
    total = int(values.get("total", 0))
    favoritos = int(values.get("favorites", 0))
    plural = "convidado" if total == 1 else "convidados"
    fav_text = f" · {favoritos} favorito" if favoritos == 1 else (f" · {favoritos} favoritos" if favoritos else "")
    row_html = (
        '<div class="family-row family-row-mobile">'
        '<div class="family-left">'
        '<div class="family-avatar">👥</div>'
        '<div class="family-copy">'
        f'<div class="family-name">{safe_family}</div>'
        f'<div class="family-sub">{total} {plural}{fav_text}</div>'
        '</div>'
        '</div>'
        '<div class="family-right">'
        f'<div class="pill">{total}</div>'
        '<div class="chevron">›</div>'
        '</div>'
        '</div>'
        '<span class="family-card-click-marker"></span>'
    )
    st.markdown(row_html, unsafe_allow_html=True)
    if st.button(f"Abrir {family_display}", key=f"open_family_{stable_key(family_display)}", use_container_width=True):
        st.session_state.selected_family = family_display
        st.session_state.editing_guest_id = None
        st.session_state.selected_guest_id = None
        st.rerun()


def render_guest_editor_panel(guest, title="Editar convidado", family_locked=False):
    if not guest:
        return

    form_key = f"edit_form_{guest['id']}"
    with st.container(border=True):
        st.markdown(
            f"""
            <div class="editor-title">{html.escape(title)}</div>
            <div class="editor-sub">Altere os dados e salve. Nada abre em outra aba.</div>
            """,
            unsafe_allow_html=True,
        )
        with st.form(form_key):
            name = st.text_input("Nome", value=guest.get("name", ""), key=f"form_name_{guest['id']}")
            family = st.text_input(
                "Família",
                value=display_family_name(guest.get("family", "Sem família")),
                key=f"form_family_{guest['id']}",
                disabled=family_locked,
            )
            col_a, col_b = st.columns(2)
            status_options = ["Pendente", "Confirmado"]
            side_options = ["Noiva", "Noivo"]
            status = col_a.selectbox(
                "Status",
                status_options,
                index=status_options.index(guest.get("status", "Pendente")) if guest.get("status", "Pendente") in status_options else 0,
                key=f"form_status_{guest['id']}",
            )
            side = col_b.selectbox(
                "Lado",
                side_options,
                index=side_options.index(guest.get("side", "Noiva")) if guest.get("side", "Noiva") in side_options else 0,
                key=f"form_side_{guest['id']}",
            )
            favorite = st.toggle("Marcar como padrinho/madrinha", value=bool(guest.get("favorite")), key=f"form_fav_{guest['id']}")
            submitted = st.form_submit_button("Salvar alterações", type="primary", use_container_width=True)

        action_cols = st.columns(2)
        if action_cols[0].button("Fechar", key=f"close_editor_{guest['id']}", use_container_width=True):
            clear_detail_params()
            st.rerun()
        if action_cols[1].button("Remover", key=f"remove_editor_{guest['id']}", use_container_width=True):
            remove_guest(guest["id"])
            clear_detail_params()
            st.rerun()

        if submitted:
            guest["name"] = name.strip()
            if not family_locked:
                guest["family"] = display_family_name(family.strip() or "Sem família")
            guest["status"] = status
            guest["side"] = side
            guest["favorite"] = favorite
            save_data()
            clear_detail_params()
            st.success("Alterações salvas.")
            st.rerun()


def render_family_editor(family):
    family = display_family_name(family)
    key = stable_key(family)
    members = [guest for guest in st.session_state.guests if family_matches(guest, family)]
    with st.container(border=True):
        st.markdown(
            f"""
            <div class="editor-title">{html.escape(family)}</div>
            <div class="editor-sub">Edite o núcleo familiar, adicione pessoas ou toque em um convidado para alterar.</div>
            """,
            unsafe_allow_html=True,
        )

        new_family_name = st.text_input(
            "Nome da família",
            value=family,
            key=f"family_rename_{key}",
            placeholder="Ex.: Família Costa",
        )
        rename_cols = st.columns(2)
        if rename_cols[0].button("Salvar nome da família", key=f"save_family_{key}", type="primary", use_container_width=True):
            rename_family(family, new_family_name)
            st.query_params.clear()
            st.query_params["page"] = "Famílias"
            st.query_params["family"] = display_family_name(new_family_name)
            st.rerun()
        if rename_cols[1].button("Fechar edição", key=f"close_family_{key}", use_container_width=True):
            clear_detail_params()
            st.rerun()

        add_key = f"family_add_guest_{key}"
        add_cols = st.columns([2.2, 1], vertical_alignment="bottom")
        add_cols[0].text_input(
            "Adicionar pessoa",
            key=add_key,
            placeholder="Nome do convidado",
        )
        if add_cols[1].button("Adicionar", key=f"add_to_family_{key}", use_container_width=True):
            add_guest_to_family(family, add_key)
            st.rerun()

        st.markdown(f"<div class='section-kicker'>{len(members)} convidados nesta família</div>", unsafe_allow_html=True)
        for guest in members:
            render_guest_card(guest, page="Famílias", family_locked=True)


def render_bottom_nav(active_page):
    nav_items = [
        ("Convidados", "👥"),
        ("Famílias", "⌂"),
        ("Padrinhos", "♡"),
    ]
    st.markdown('<div id="mobile-bottom-nav-marker"></div>', unsafe_allow_html=True)
    nav_cols = st.columns(3)
    for col, (label, icon) in zip(nav_cols, nav_items):
        button_type = "primary" if label == active_page else "secondary"
        if col.button(
            f"{icon}\n{label}",
            key=f"mobile_nav_{label}",
            type=button_type,
            use_container_width=True,
        ):
            set_page(label)
            st.rerun()


# -----------------------------
# Configuração visual
# -----------------------------
st.set_page_config(page_title="Lista do casamento", layout="wide")

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700&family=Inter:wght@400;500;600;700&display=swap');

    #MainMenu, footer, header[data-testid="stHeader"], [data-testid="stToolbar"], [data-testid="stDecoration"], [data-testid="stStatusWidget"] {
        display: none !important;
        visibility: hidden !important;
        height: 0 !important;
    }
    :root {
        --rose: #c77d75;
        --rose-dark: #b86d66;
        --rose-soft: #e2aaa1;
        --rose-light: #f8eeee;
        --ink: #171827;
        --muted: #858891;
        --line: #eadfda;
        --card: rgba(255, 255, 255, 0.92);
        --wash: #fbf6f2;
        --shadow: 0 18px 46px rgba(72, 41, 35, 0.08);
    }
    html, body, [data-testid="stAppViewContainer"] {
        background:
            radial-gradient(circle at 16% 5%, rgba(199, 125, 117, 0.12), transparent 32%),
            radial-gradient(circle at 95% 0%, rgba(199, 125, 117, 0.08), transparent 28%),
            linear-gradient(180deg, #fffbf8 0%, #f9f4ef 100%);
    }
    .block-container {
        padding: 2.2rem 2.1rem 5.8rem;
        max-width: 1180px;
    }
    section[data-testid="stSidebar"] {
        background: rgba(255, 251, 248, 0.96);
        border-right: 1px solid var(--line);
    }
    h1, .app-title {
        color: var(--rose);
        font-family: 'Playfair Display', Georgia, serif;
        font-size: clamp(2.55rem, 7vw, 4.35rem);
        line-height: 0.95;
        margin: 0;
    }
    h2, h3, p, label, div, span, input, button, a {
        font-family: 'Inter', system-ui, sans-serif;
    }
    .app-header {
        display: flex;
        align-items: flex-start;
        justify-content: space-between;
        gap: 1rem;
        margin-bottom: 1.15rem;
    }
    .app-subtitle,
    [data-testid="stCaptionContainer"] {
        color: var(--muted);
        font-size: 1.03rem;
        margin-top: 0.55rem;
    }
    .header-icon {
        color: var(--rose);
        font-size: 1.65rem;
        line-height: 1;
        padding-top: 0.45rem;
        opacity: 0.95;
    }
    input, textarea, [data-baseweb="select"] > div {
        border-radius: 14px !important;
    }
    .stButton button,
    .stDownloadButton button,
    .stFormSubmitButton button {
        border-radius: 14px;
        min-height: 44px;
        border-color: var(--line);
        color: var(--rose-dark);
        background: rgba(255, 255, 255, 0.9);
        font-weight: 650;
    }
    .stButton button[kind="primary"],
    .stDownloadButton button,
    .stFormSubmitButton button[kind="primary"] {
        background: linear-gradient(135deg, #dca8a0, #c8746f);
        border: 0;
        color: #fff;
    }
    div[data-testid="stVerticalBlockBorderWrapper"] {
        border-color: var(--line);
        border-radius: 18px;
        background: var(--card);
        box-shadow: 0 10px 25px rgba(72, 41, 35, 0.06);
    }
    .hero-card, .soft-card, .stats-card {
        border: 1px solid var(--line);
        background: var(--card);
        border-radius: 20px;
        box-shadow: var(--shadow);
    }
    .hero-card {
        padding: 1.15rem;
        margin: 1.25rem 0 1.15rem;
    }
    .soft-card {
        padding: 1.2rem 1.3rem;
        margin: 1.1rem 0 1.2rem;
    }
    .stats-card {
        display: flex;
        align-items: center;
        justify-content: space-around;
        gap: 1rem;
        min-height: 112px;
        padding: 1.1rem 1.2rem;
        margin: 1.1rem 0 1.2rem;
    }
    .stats-item { display: flex; align-items: center; gap: 1rem; min-width: 0; }
    .stats-icon, .family-avatar, .guest-avatar {
        width: 56px;
        height: 56px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        background: var(--rose-light);
        color: var(--rose);
        font-size: 1.35rem;
        flex: 0 0 auto;
    }
    .stats-number {
        color: var(--ink);
        font-family: 'Playfair Display', Georgia, serif;
        font-size: 2.25rem;
        font-weight: 700;
        line-height: 1;
    }
    .stats-label { color: var(--muted); font-size: 1rem; margin-top: 0.35rem; }
    .stats-divider { width: 1px; align-self: stretch; background: var(--line); }
    .mobile-list-top {
        display: flex;
        align-items: center;
        justify-content: space-between;
        color: var(--ink);
        margin: 1.25rem 0 .65rem;
        gap: 1rem;
    }
    .mobile-list-count {
        display: inline-flex;
        align-items: center;
        gap: .55rem;
        font-weight: 700;
        color: var(--ink);
        font-size: 1.08rem;
    }
    .mobile-list-count span:first-child { color: var(--rose); }
    .mobile-order { color: var(--muted); font-weight: 500; }
    .guest-card-link {
        text-decoration: none !important;
        color: inherit !important;
        display: block;
    }
    .guest-card, .family-row {
        display: flex;
        align-items: center;
        border: 1px solid var(--line);
        background: var(--card);
        border-radius: 20px;
        padding: 1rem 1.05rem;
        margin: .72rem 0;
        box-shadow: 0 10px 25px rgba(72, 41, 35, 0.05);
    }
    .guest-card:hover, .family-row:hover {
        transform: translateY(-1px);
        transition: transform .15s ease;
        border-color: rgba(199, 125, 117, .35);
    }
    .guest-card-open {
        border-color: rgba(199, 125, 117, .55);
        background: linear-gradient(135deg, rgba(255,255,255,.96), rgba(253,243,241,.96));
        box-shadow: 0 12px 28px rgba(199, 125, 117, .10);
    }

    /* Botão invisível sobre o card: mantém o visual HTML e deixa o card clicável pelo Streamlit. */
    div[data-testid="stElementContainer"]:has(.guest-card-click-marker) + div[data-testid="stElementContainer"],
    div[data-testid="stElementContainer"]:has(.family-card-click-marker) + div[data-testid="stElementContainer"] {
        margin-top: -92px !important;
        height: 92px !important;
        position: relative !important;
        z-index: 20 !important;
        opacity: 0 !important;
    }
    div[data-testid="stElementContainer"]:has(.guest-card-click-marker) + div[data-testid="stElementContainer"] .stButton,
    div[data-testid="stElementContainer"]:has(.family-card-click-marker) + div[data-testid="stElementContainer"] .stButton,
    div[data-testid="stElementContainer"]:has(.guest-card-click-marker) + div[data-testid="stElementContainer"] .stButton button,
    div[data-testid="stElementContainer"]:has(.family-card-click-marker) + div[data-testid="stElementContainer"] .stButton button {
        width: 100% !important;
        height: 92px !important;
        min-height: 92px !important;
        cursor: pointer !important;
    }
    .guest-copy, .family-copy { min-width: 0; flex: 1; }
    .guest-name, .family-name {
        color: var(--ink);
        font-family: 'Playfair Display', Georgia, serif;
        font-weight: 700;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    .guest-name { font-size: 1.35rem; }
    .family-name { font-size: 1.45rem; }
    .guest-family, .family-sub {
        color: var(--muted);
        margin-top: .22rem;
        font-size: 1rem;
    }
    .guest-chevron, .chevron {
        color: var(--muted);
        font-size: 2rem;
        line-height: 1;
        padding-left: .75rem;
    }
    .guest-avatar { margin-right: 1rem; }
    .family-left, .family-right { display: flex; align-items: center; gap: 1rem; min-width: 0; }
    .family-left { flex: 1; }
    .family-right { flex: 0 0 auto; }
    .pill {
        background: linear-gradient(135deg, #dca8a0, #c8746f);
        color: white;
        border-radius: 13px;
        padding: .48rem .75rem;
        font-weight: 700;
        min-width: 2.65rem;
        text-align: center;
        box-shadow: 0 8px 16px rgba(199, 125, 117, .22);
    }
    .mini-badge {
        background: #fbecf5;
        color: #9f5f87;
        border: 1px solid #efd4e4;
        border-radius: 999px;
        padding: .28rem .55rem;
        font-size: .78rem;
        font-weight: 700;
        margin-left: .7rem;
    }
    .section-kicker {
        color: var(--muted);
        font-weight: 650;
        margin: .35rem 0 .8rem;
    }
    .search-helper {
        display: flex;
        align-items: center;
        justify-content: center;
        height: 44px;
        border: 1px solid var(--line);
        border-radius: 14px;
        background: rgba(255, 255, 255, 0.72);
        color: var(--rose-dark);
        font-size: 1.25rem;
    }
    .editor-title {
        color: var(--ink);
        font-family: 'Playfair Display', Georgia, serif;
        font-size: 1.72rem;
        font-weight: 700;
        margin-bottom: .25rem;
    }
    .editor-sub {
        color: var(--muted);
        font-size: .96rem;
        margin-bottom: 1rem;
    }
    .padrinhos-tabs {
        display: grid;
        grid-template-columns: 1fr 1fr;
        border: 1px solid var(--line);
        border-radius: 14px;
        overflow: hidden;
        margin: 1.3rem 0 1.2rem;
        background: rgba(255,255,255,.7);
    }
    .padrinhos-tab {
        padding: .95rem 1rem;
        text-align: center;
        color: var(--muted);
        font-weight: 650;
    }
    .padrinhos-tab.active {
        color: white;
        background: linear-gradient(135deg, #dca8a0, #c8746f);
    }
    .padrinhos-hero {
        position: relative;
        overflow: hidden;
        border: 1px solid var(--line);
        background: linear-gradient(135deg, rgba(255,255,255,.95), rgba(248,238,238,.9));
        border-radius: 20px;
        padding: 1.4rem 1.25rem;
        margin: 1rem 0 1.25rem;
        box-shadow: var(--shadow);
        display: flex;
        align-items: center;
        gap: 1.2rem;
    }
    .padrinhos-heart {
        width: 76px;
        height: 76px;
        border-radius: 50%;
        background: white;
        display: flex;
        align-items: center;
        justify-content: center;
        color: var(--rose);
        font-size: 2rem;
        box-shadow: 0 10px 24px rgba(72, 41, 35, .08);
    }
    #mobile-bottom-nav-marker { display: none; }

    @media (max-width: 720px) {
        .block-container {
            padding: 1.65rem 1.05rem 7rem;
            max-width: 480px;
        }
        section[data-testid="stSidebar"] { display: none; }
        .app-header { margin: .25rem 0 1.25rem; }
        .app-title { font-size: 3.05rem; }
        .app-subtitle { font-size: 1rem; line-height: 1.35; }
        .header-icon { font-size: 1.45rem; }
        .hero-card { padding: .95rem; margin-top: 1rem; }
        .stats-card { min-height: 104px; justify-content: space-between; padding: 1rem; }
        .stats-item { gap: .8rem; flex: 1 1 0; }
        .stats-icon, .family-avatar, .guest-avatar { width: 50px; height: 50px; font-size: 1.18rem; }
        .stats-number { font-size: 2.05rem; }
        .stats-label { font-size: .94rem; }
        .guest-card, .family-row { border-radius: 18px; padding: .92rem .88rem; margin: .65rem 0; }
        .guest-avatar { margin-right: .9rem; }
        .guest-name { font-size: 1.24rem; }
        .family-name { font-size: 1.26rem; }
        .guest-family, .family-sub { font-size: .95rem; }
        .guest-chevron, .chevron { font-size: 1.8rem; padding-left: .45rem; }
        .pill { min-width: 2.45rem; padding: .46rem .68rem; }
        .mini-badge { display: none; }
        .mobile-list-top { margin-top: 1.05rem; }
        .padrinhos-hero { padding: 1.15rem 1rem; gap: 1rem; }
        .padrinhos-heart { width: 62px; height: 62px; font-size: 1.7rem; }
        .editor-title { font-size: 1.5rem; }

        div[data-testid="stElementContainer"]:has(#mobile-bottom-nav-marker) + div[data-testid="stHorizontalBlock"],
        .element-container:has(#mobile-bottom-nav-marker) + div[data-testid="stHorizontalBlock"] {
            position: fixed;
            left: 0;
            right: 0;
            bottom: 0;
            display: grid !important;
            grid-template-columns: repeat(3, 1fr);
            gap: 0;
            background: rgba(255,255,255,0.94);
            backdrop-filter: blur(18px);
            border-top: 1px solid var(--line);
            padding: .45rem .65rem calc(.55rem + env(safe-area-inset-bottom));
            z-index: 1000;
            box-shadow: 0 -14px 32px rgba(72, 41, 35, .08);
        }
        div[data-testid="stElementContainer"]:has(#mobile-bottom-nav-marker) + div[data-testid="stHorizontalBlock"] .stButton button,
        .element-container:has(#mobile-bottom-nav-marker) + div[data-testid="stHorizontalBlock"] .stButton button {
            min-height: 58px;
            border: 0;
            background: transparent;
            box-shadow: none;
            color: var(--muted);
            font-weight: 650;
            white-space: pre-line;
            line-height: 1.25;
            justify-content: center;
            border-radius: 16px;
        }
        div[data-testid="stElementContainer"]:has(#mobile-bottom-nav-marker) + div[data-testid="stHorizontalBlock"] .stButton button[kind="primary"],
        .element-container:has(#mobile-bottom-nav-marker) + div[data-testid="stHorizontalBlock"] .stButton button[kind="primary"] {
            color: var(--rose);
            background: transparent;
        }
        div[data-testid="stElementContainer"]:has(#mobile-bottom-nav-marker) + div[data-testid="stHorizontalBlock"] .stButton button[kind="primary"]::before,
        .element-container:has(#mobile-bottom-nav-marker) + div[data-testid="stHorizontalBlock"] .stButton button[kind="primary"]::before {
            content: "";
            position: absolute;
            top: 0;
            width: 58px;
            height: 4px;
            border-radius: 999px;
            background: var(--rose);
        }
    }

    @media (min-width: 721px) {
        div[data-testid="stElementContainer"]:has(#mobile-bottom-nav-marker),
        .element-container:has(#mobile-bottom-nav-marker),
        div[data-testid="stElementContainer"]:has(#mobile-bottom-nav-marker) + div[data-testid="stHorizontalBlock"],
        .element-container:has(#mobile-bottom-nav-marker) + div[data-testid="stHorizontalBlock"] {
            display: none !important;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# -----------------------------
# Estado inicial
# -----------------------------
if "guests" not in st.session_state:
    st.session_state.guests = load_data()["guests"]
if "page" not in st.session_state:
    st.session_state.page = "Convidados"
if "side_filter" not in st.session_state:
    st.session_state.side_filter = "Todos"
if "selected_family" not in st.session_state:
    st.session_state.selected_family = None
if "selected_guest_id" not in st.session_state:
    st.session_state.selected_guest_id = None
if "editing_guest_id" not in st.session_state:
    st.session_state.editing_guest_id = None

query_page = qp_value("page")
if query_page in ["Convidados", "Famílias", "Padrinhos"]:
    st.session_state.page = query_page

query_edit = qp_value("edit")
if query_edit and find_guest(query_edit):
    st.session_state.editing_guest_id = query_edit
    st.session_state.selected_guest_id = query_edit

query_family = qp_value("family")
if query_family and st.session_state.page == "Famílias":
    st.session_state.selected_family = display_family_name(query_family)


# -----------------------------
# Sidebar desktop
# -----------------------------
with st.sidebar:
    st.markdown(
        """
        <div style="display:flex; align-items:center; gap:.7rem; margin-bottom:2rem;">
            <div style="width:52px;height:52px;border:1px solid #eadfda;border-radius:50%;display:flex;align-items:center;justify-content:center;color:#c77d75;font-size:26px;">♡</div>
            <div style="font-family:'Playfair Display', Georgia, serif;color:#c77d75;font-size:1.45rem;font-weight:700;">Meu Casamento</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    for label in ["Convidados", "Famílias", "Padrinhos"]:
        st.button(
            label,
            use_container_width=True,
            type="primary" if st.session_state.page == label else "secondary",
            on_click=set_page,
            args=(label,),
        )
    st.markdown("<div style='height:36vh'></div>", unsafe_allow_html=True)
    st.caption("Tudo salvo automaticamente.")


total_guests = len([g for g in st.session_state.guests if g["name"].strip()])
favorite_count = len([g for g in st.session_state.guests if g["name"].strip() and g["favorite"]])
families = family_summary()
family_count = len(families)

page_subtitle = {
    "Convidados": "Organize sua lista do casamento",
    "Famílias": "Organize os convidados por núcleo familiar",
    "Padrinhos": "Lista especial do casamento",
}[st.session_state.page]

render_header(st.session_state.page, page_subtitle, {"Convidados": "👥", "Famílias": "👥", "Padrinhos": "♡"}[st.session_state.page])


# -----------------------------
# Página Convidados
# -----------------------------
if st.session_state.page == "Convidados":
    st.markdown('<div class="hero-card">', unsafe_allow_html=True)
    add_cols = st.columns([2.5, 1.4, 1, 1.1], vertical_alignment="bottom")
    add_cols[0].text_input(
        "Novo convidado",
        key="new_guest_name",
        placeholder="Digite o nome do convidado",
        label_visibility="collapsed",
    )
    family_options = ["Sem família"] + sorted(families.keys())
    add_cols[1].selectbox(
        "Família",
        family_options,
        key="new_guest_family",
        label_visibility="collapsed",
    )
    add_cols[2].selectbox(
        "Lado",
        ["Noiva", "Noivo"],
        key="new_guest_side",
        label_visibility="collapsed",
    )
    add_cols[3].button(
        "+ Adicionar",
        type="primary",
        use_container_width=True,
        on_click=add_guest_from_inputs,
    )
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown(
        f"""
        <div class="mobile-list-top">
            <div class="mobile-list-count"><span>👥</span> {total_guests} convidados</div>
            <div class="mobile-order">Ordenar⌄</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    filter_cols = st.columns([1, 1, 1, 2])
    for label, col in zip(["Todos", "Noiva", "Noivo"], filter_cols[:3]):
        if col.button(
            label,
            use_container_width=True,
            type="primary" if st.session_state.side_filter == label else "secondary",
        ):
            st.session_state.side_filter = label
            st.rerun()

    with st.expander("Importar / baixar lista", expanded=False):
        tool_cols = st.columns([1, 1])
        if tool_cols[0].button("Importar Pasta1.xlsx", use_container_width=True):
            if DEFAULT_XLSX.exists():
                import_xlsx(DEFAULT_XLSX)
                st.rerun()
            else:
                st.error("Não encontrei a planilha no caminho configurado.")
        tool_cols[1].download_button(
            "Baixar CSV",
            data=csv_bytes(),
            file_name="lista_casamento.csv",
            mime="text/csv",
            use_container_width=True,
        )

    shown_guests = [
        guest
        for guest in st.session_state.guests
        if guest["name"].strip()
        and (st.session_state.side_filter == "Todos" or guest.get("side", "Noiva") == st.session_state.side_filter)
    ]
    for guest in shown_guests:
        render_guest_card(guest, page="Convidados")


# -----------------------------
# Página Famílias
# -----------------------------
elif st.session_state.page == "Famílias":
    render_family_stats(family_count, total_guests)
    search_cols = st.columns([5, .7], vertical_alignment="center")
    search = search_cols[0].text_input(
        "Buscar família",
        placeholder="Buscar família",
        label_visibility="collapsed",
    ).strip().lower()
    search_cols[1].markdown('<div class="search-helper">☰</div>', unsafe_allow_html=True)

    if st.session_state.selected_family:
        render_family_editor(st.session_state.selected_family)

    if not families:
        st.info("Nenhum convidado cadastrado.")
    else:
        visible_families = []
        for family, values in sorted(families.items()):
            if search and search not in family.lower():
                continue
            visible_families.append((family, values))

        for family, values in visible_families:
            render_family_row(family, values)

        if not visible_families:
            st.info("Nenhuma família encontrada com esse filtro.")


# -----------------------------
# Página Padrinhos
# -----------------------------
else:
    st.markdown(
        """
        <div class="padrinhos-tabs">
            <div class="padrinhos-tab">Noivos</div>
            <div class="padrinhos-tab active">Madrinhas & Padrinhos</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        f"""
        <div class="padrinhos-hero">
            <div class="padrinhos-heart">♡</div>
            <div>
                <div class="family-name">{favorite_count} padrinhos</div>
                <div class="family-sub">Pessoas especiais que caminham com vocês nesse dia inesquecível.</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if st.button("+ Adicionar padrinho", type="primary", use_container_width=True):
        guest = new_guest("Novo padrinho", "Padrinhos", True)
        st.session_state.guests.insert(0, guest)
        save_data()
        st.session_state.editing_guest_id = guest["id"]
        st.session_state.selected_guest_id = guest["id"]
        st.rerun()

    favorites = [guest for guest in st.session_state.guests if guest["name"].strip() and guest["favorite"]]
    if not favorites:
        st.info("Nenhum padrinho ou madrinha marcado ainda.")
    for guest in favorites:
        render_guest_card(guest, page="Padrinhos")


render_bottom_nav(st.session_state.page)
